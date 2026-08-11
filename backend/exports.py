"""
Export helpers matching Solvi's "Share and Collaborate" export formats:
GeoTIFF, SHP (via GeoJSON->shapefile), Excel, and a simple PDF summary report.
"""
import io
import zipfile
from pathlib import Path

import numpy as np
import rasterio
from rasterio.transform import Affine
import geopandas as gpd
from shapely.geometry import shape
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def index_array_to_geotiff_bytes(index_array: np.ndarray, transform, crs) -> bytes:
    """Writes a single-band float32 GeoTIFF (e.g. an NDVI raster) to bytes."""
    buf = io.BytesIO()
    profile = {
        "driver": "GTiff",
        "dtype": "float32",
        "nodata": np.nan,
        "width": index_array.shape[1],
        "height": index_array.shape[0],
        "count": 1,
        "crs": crs,
        "transform": transform,
        "compress": "lzw",
    }
    with rasterio.io.MemoryFile() as memfile:
        with memfile.open(**profile) as dst:
            dst.write(index_array.astype(np.float32), 1)
        buf.write(memfile.read())
    return buf.getvalue()


def geometries_to_shapefile_zip_bytes(features: list[dict], crs: str = "EPSG:4326",
                                       layer_name: str = "export") -> bytes:
    """
    features: list of {geometry: GeoJSON dict, **attributes}
    Shapefiles are multi-file formats (.shp/.shx/.dbf/.prj) — this writes
    them to a temp dir and zips the set, which is what Solvi's "export to SHP"
    effectively hands you too.
    """
    rows = []
    for f in features:
        row = {k: v for k, v in f.items() if k != "geometry"}
        row["geometry"] = shape(f["geometry"])
        rows.append(row)

    gdf = gpd.GeoDataFrame(rows, crs=crs)

    import tempfile
    with tempfile.TemporaryDirectory() as tmpdir:
        shp_path = Path(tmpdir) / f"{layer_name}.shp"
        gdf.to_file(shp_path, driver="ESRI Shapefile")

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for ext in [".shp", ".shx", ".dbf", ".prj", ".cpg"]:
                p = Path(tmpdir) / f"{layer_name}{ext}"
                if p.exists():
                    zf.write(p, arcname=f"{layer_name}{ext}")
        return buf.getvalue()


def plot_stats_to_excel_bytes(plot_stats: list[dict], sheet_title: str = "Plot Statistics") -> bytes:
    """
    plot_stats: list of dicts, e.g.
      {plot_id, row, col, area_m2, mean, min, max, std, pixel_count}
    Produces a formatted Excel workbook (color-scaled mean column), matching
    Solvi's "extensive statistics for each plot ... exports to Excel" feature.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    if not plot_stats:
        ws.append(["No plot statistics available"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(plot_stats[0].keys())
    ws.append(headers)
    header_fill = PatternFill(start_color="2C7A45", end_color="2C7A45", fill_type="solid")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row_data in plot_stats:
        ws.append([row_data.get(h) for h in headers])

    # color-scale the 'mean' column if present (low=red, high=green) — quick visual QA
    if "mean" in headers:
        mean_col_idx = headers.index("mean") + 1
        mean_col_letter = get_column_letter(mean_col_idx)
        values = [r.get("mean") for r in plot_stats if r.get("mean") is not None]
        if values:
            vmin, vmax = min(values), max(values)
            for i, row_data in enumerate(plot_stats, start=2):
                v = row_data.get("mean")
                if v is None or vmax == vmin:
                    continue
                t = (v - vmin) / (vmax - vmin)
                r = int(215 + (26 - 215) * t)
                g = int(48 + (152 - 48) * t)
                b = int(39 + (80 - 39) * t)
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                ws[f"{mean_col_letter}{i}"].fill = PatternFill(
                    start_color=hex_color, end_color=hex_color, fill_type="solid"
                )

    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(h)) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_report_bytes(farm_name: str, index_label: str, overall_stats: dict,
                            plot_count: int, overlay_png_bytes: bytes = None) -> bytes:
    """A short one-page PDF summary: farm name, index, field-level stats, plot count,
    and (optionally) the index overlay image — matching Solvi's PDF export option."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    c.setFillColor(colors.HexColor("#1f5c33"))
    c.rect(0, height - 30 * mm, width, 30 * mm, fill=True, stroke=False)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(15 * mm, height - 18 * mm, "FarmScan Field Report")
    c.setFont("Helvetica", 11)
    c.drawString(15 * mm, height - 25 * mm, farm_name)

    y = height - 42 * mm
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(15 * mm, y, f"Index: {index_label}")
    y -= 8 * mm

    c.setFont("Helvetica", 11)
    for label, key in [("Mean", "mean"), ("Min", "min"), ("Max", "max"),
                        ("Std dev", "std"), ("Pixels sampled", "pixel_count")]:
        val = overall_stats.get(key)
        val_str = f"{val:.3f}" if isinstance(val, float) else str(val)
        c.drawString(18 * mm, y, f"{label}: {val_str}")
        y -= 6 * mm

    c.drawString(18 * mm, y, f"Plots analyzed: {plot_count}")
    y -= 10 * mm

    if overlay_png_bytes:
        from reportlab.lib.utils import ImageReader
        img = ImageReader(io.BytesIO(overlay_png_bytes))
        img_w = 120 * mm
        img_h = 120 * mm
        c.drawImage(img, 15 * mm, y - img_h, width=img_w, height=img_h,
                    preserveAspectRatio=True, mask="auto")

    c.showPage()
    c.save()
    return buf.getvalue()
